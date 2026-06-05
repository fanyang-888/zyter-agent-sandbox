"""
Knowledge base setup script.

Run this once before starting the app (and after adding new source docs):
  python scripts/setup_kb.py

What it does:
  1. Reads all .txt, .pdf, .md files from ./source_docs/
  2. Chunks them by collection tag in the filename
  3. Embeds and stores in Chroma

Collection routing (by filename prefix):
  universal_*    -> "universal" collection  (all agents)
  rfp_*          -> "rfp_archive" collection (RFP agent only)
  ci_*           -> "competitive_intel" collection (CI agent only)
  (no prefix)    -> "universal" by default

Usage:
  Add a doc to source_docs/ with the right prefix, re-run this script.
  Chroma will skip already-indexed chunks (content hash dedup).
"""

import os
import sys
import hashlib
from pathlib import Path

from dotenv import load_dotenv
try:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
except ImportError:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain.schema import Document
from langchain_community.vectorstores import Chroma
from langchain_openai import OpenAIEmbeddings

load_dotenv()

SOURCE_DOCS_DIR = Path(os.getenv("SOURCE_DOCS_DIR", "./source_docs"))
CHROMA_PERSIST_DIR = os.getenv("CHROMA_PERSIST_DIR", "./chroma_db")
EMBEDDING_MODEL = "text-embedding-3-large"

COLLECTION_MAP = {
    "universal_": "universal",
    "rfp_": "rfp_archive",
    "ci_": "competitive_intel",
}

splitter = RecursiveCharacterTextSplitter(
    chunk_size=800,
    chunk_overlap=100,
    separators=["\n\n", "\n", ". ", " ", ""],
)


def collection_for_file(filename: str) -> str:
    for prefix, collection in COLLECTION_MAP.items():
        if filename.lower().startswith(prefix):
            return collection
    return "universal"


def read_file(path: Path) -> str:
    suffix = path.suffix.lower()

    if suffix in (".txt", ".md"):
        return path.read_text(encoding="utf-8", errors="ignore")

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            print(f"  [SKIP] pypdf not installed, cannot read {path.name}")
            return ""

    if suffix in (".docx",):
        try:
            import zipfile, xml.etree.ElementTree as ET
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            with zipfile.ZipFile(str(path)) as z:
                with z.open("word/document.xml") as f:
                    tree = ET.parse(f)
            paragraphs = tree.getroot().findall(".//w:p", ns)
            lines = []
            for p in paragraphs:
                text = "".join(r.text or "" for r in p.findall(".//w:t", ns))
                if text.strip():
                    lines.append(text)
            return "\n".join(lines)
        except Exception as e:
            print(f"  [SKIP] could not read docx {path.name}: {e}")
            return ""

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    line = " | ".join(cells).strip(" |")
                    if line.replace("|", "").strip():
                        parts.append(line)
            return "\n".join(parts)
        except ImportError:
            print(f"  [SKIP] openpyxl not installed — run: pip install openpyxl")
            return ""
        except Exception as e:
            print(f"  [SKIP] could not read xlsx {path.name}: {e}")
            return ""

    print(f"  [SKIP] unsupported file type: {path.name}")
    return ""


DOC_TYPE_MAP = {
    # filename keyword      -> doc_type tag
    "product_overview":     "product_spec",
    "battle_card":          "battle_card",
    "battle_cards":         "battle_card",
    "rfp":                  "rfp_response",
    "playbook":             "sales_playbook",
    "win_loss":             "win_loss",
    "competitor":           "competitor_note",
    "handbook":             "internal_handbook",
    "release_notes":        "release_notes",
    "release":              "release_notes",
    "training":             "training_doc",
    "onboarding":           "onboarding_doc",
    "rrtraining":           "training_doc",
    "releaseconsideration": "release_notes",
}

TEAM_MAP = {
    "sales":       "sales",
    "gtm":         "sales",
    "engineering": "engineering",
    "product":     "product",
    "operations":  "operations",
    "marketing":   "marketing",
    "cs":          "customer_success",
    "support":     "customer_success",
}


def _doc_type_for_file(filename: str) -> str:
    lower = filename.lower()
    for keyword, doc_type in DOC_TYPE_MAP.items():
        if keyword in lower:
            return doc_type
    return "general"


def _team_for_file(filename: str) -> str:
    lower = filename.lower()
    for keyword, team in TEAM_MAP.items():
        if keyword in lower:
            return team
    return "all"


def _freshness_from_filename(filename: str) -> str:
    """
    Extracts YYYY-MM from filename if present (e.g. battle_cards_2026-05.txt).
    Falls back to current month so freshness tracking starts immediately.
    """
    import re
    match = re.search(r"(\d{4}-\d{2})", filename)
    if match:
        return match.group(1)
    from datetime import date
    return date.today().strftime("%Y-%m")


def chunk_document(path: Path) -> list[Document]:
    text = read_file(path)
    if not text.strip():
        return []

    chunks = splitter.split_text(text)
    doc_type  = _doc_type_for_file(path.name)
    team      = _team_for_file(path.name)
    freshness = _freshness_from_filename(path.name)

    return [
        Document(
            page_content=chunk,
            metadata={
                "source":     path.name,
                "collection": collection_for_file(path.name),
                "doc_type":   doc_type,
                "team":       team,
                "freshness":  freshness,
                "chunk_hash": hashlib.md5(chunk.encode()).hexdigest(),
            },
        )
        for chunk in chunks
    ]


def ingest():
    if not SOURCE_DOCS_DIR.exists():
        print(f"source_docs/ not found at {SOURCE_DOCS_DIR}. Creating it...")
        SOURCE_DOCS_DIR.mkdir(parents=True)
        print("Add .txt, .pdf, or .md files to source_docs/ and re-run.")
        sys.exit(0)

    embeddings = OpenAIEmbeddings(model=EMBEDDING_MODEL)
    vectorstores: dict[str, Chroma] = {}

    files = list(SOURCE_DOCS_DIR.glob("*"))
    readable = [f for f in files if f.suffix.lower() in (".txt", ".pdf", ".md", ".docx", ".xlsx")]

    if not readable:
        print("No readable files found in source_docs/. Add .txt, .pdf, or .md files.")
        sys.exit(0)

    print(f"Found {len(readable)} file(s) to ingest:\n")

    all_docs: dict[str, list[Document]] = {}

    for path in readable:
        collection = collection_for_file(path.name)
        chunks = chunk_document(path)
        print(f"  {path.name} → collection='{collection}', {len(chunks)} chunks")
        all_docs.setdefault(collection, []).extend(chunks)

    print("\nEmbedding and storing in Chroma...")
    for collection_name, docs in all_docs.items():
        vs = Chroma(
            collection_name=collection_name,
            embedding_function=embeddings,
            persist_directory=CHROMA_PERSIST_DIR,
        )
        # Deduplicate by chunk_hash before adding
        existing_ids = set(vs.get()["ids"]) if vs.get()["ids"] else set()
        new_docs = [d for d in docs if d.metadata["chunk_hash"] not in existing_ids]

        if new_docs:
            vs.add_documents(new_docs, ids=[d.metadata["chunk_hash"] for d in new_docs])
            print(f"  [{collection_name}] added {len(new_docs)} new chunks "
                  f"({len(docs) - len(new_docs)} already existed)")
        else:
            print(f"  [{collection_name}] all {len(docs)} chunks already indexed, skipping")

    print("\nKnowledge base ready.")
    print(f"Chroma stored at: {CHROMA_PERSIST_DIR}")


if __name__ == "__main__":
    ingest()
